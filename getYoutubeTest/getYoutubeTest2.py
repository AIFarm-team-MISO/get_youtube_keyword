"""
키워드를 입력하면 유튜브 구독자 5000명 미만 채널 중 조회수 10만 이상이고 지금도 계속해서 조회수가 상승하는영상의 링크를 가져와 엑셀 파일로 저장하는 코드를 알려주세요.
-> If you enter a keyword, please tell me the code to get a link to a video that has more than 100,000 views among channels with less than 5,000 YouTube subscribers and continues to increase in number of views and save it as an Excel file.

"""

import pytube
import requests
import openpyxl
from bs4 import BeautifulSoup

keyword = input("Enter keyword: ")
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['Title', 'Views', 'Channel Name', 'Subscribers', 'Video URL'])

url = f'https://www.youtube.com/results?search_query={keyword}'
html = requests.get(url).text
soup = BeautifulSoup(html, 'html.parser')

for vid in soup.findAll(attrs={'class':'yt-uix-tile-link'}):
    video_url = 'https://www.youtube.com' + vid['href']
    try:
        yt = pytube.YouTube(video_url)
        if yt.views > 100000:
            channel_url = 'https://www.youtube.com' + vid['href'].split('/')[1]
            html = requests.get(channel_url).text
            soup = BeautifulSoup(html, 'html.parser')
            sub_count = soup.find('span', {'class': 'yt-subscription-button-subscriber-count-branded-horizontal yt-subscriber-count'}).text.strip()
            sub_count = sub_count.replace(' subscribers', '').replace(',', '')
            if sub_count and int(sub_count) < 5000 and not yt.streams.filter(live=True):
                sheet.append([yt.title, yt.views, yt.author, sub_count, video_url])
                print(f"Added video: {yt.title}")
    except (pytube.exceptions.VideoUnavailable, pytube.exceptions.RegexMatchError):
        print(f"Video at {video_url} is unavailable or does not meet the criteria.")
    except Exception as e:
        print(f"An error occurred while processing {video_url}: {e}")

if sheet.max_row > 1:
    wb.save(f'{keyword} results.xlsx')
    print(f"{sheet.max_row-1} video(s) saved to {keyword} results.xlsx")
else:
    print(f"No video found for {keyword} that meets the criteria.")

